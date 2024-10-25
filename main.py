from pypdf import PdfReader
from pathlib import Path
from PIL import Image
import pytesseract
import os
from docx import Document
from openpyxl import load_workbook
import pandas as pd

_EXCEPTION_MESSAGE_FILE_NOT_FOUND = ("The file that was passed has not been found. Please check the file exists or "
                                     "that the file path is not broken and try again.")


# This function could further be improved to support PDF files with multiple columns and different orientations
def extract_pdf(pdf_path):
    """ This function extracts the text from a PDF file and returns the content in a string.

    Args:
        pdf_path(str): Path to the PDF file that is to be converted"""
    name = Path(pdf_path).name

    try:
        content = PdfReader(pdf_path)
        text = ""
        for page in content.pages:
            text += page.extract_text() + "\n"

        print(f"Successfully extracted text from {name}!")
        return text

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


# This function could further be improved by pre-processing images in order for the text to be more accurately recognised
def extract_image(img_path):
    """This function extracts the text from an image file (.jpeg, .png, .tiff, .gif, .bmp etc.) and
    returns the content in a string.

     Args:
        img_path(str): Path to the Image file that is to be converted."""

    name = Path(img_path).name

    try:
        image = Image.open(img_path)
        text = pytesseract.image_to_string(image)

        print(f"Successfully extracted text from {name}!")
        return text

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


# This function can be further improved to be able to detect better headers, hyperlinks etc
def extract_docx(word_file_path):
    """This function extracts the text from a Word file (.docx) and
       returns the content in a string.

        Args:
           word_file_path(str): Path to the Word file that is to be converted."""

    name = Path(word_file_path).name

    try:
        text = ""
        document = Document(word_file_path)

        for paragraph in document.paragraphs:
            text += paragraph.text + "\n"

            print(f"Successfully extracted text from {name}!")
            return text

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


def extract_txt(txt_file_path):
    """This function extracts the text from a text file (.txt) and
       returns the content in a string.

        Args:
           txt_file_path(str): Path to the Word file that is to be converted."""

    name = Path(txt_file_path).name

    try:
        with open(txt_file_path, "r") as file:
            text = file.readlines()

            print(f"Successfully extracted text from {name}!")
            return text

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


def extract_excel(excel_path):
    """ This function extracts the text from an Excel file (.xlsx, .xlsm, .xltx, .xltm) and returns the content
    as list of lists of strings. Each row is represented by a list of strings that contain the cell values.

        Args:
            excel_path(str): Path to the Excel file that is to be converted"""

    name = Path(excel_path).name

    try:
        workbook = load_workbook(excel_path)
        sheets = workbook.sheetnames

        data = []

        for sheet in sheets:
            current_sheet = workbook[sheet]
            for row in current_sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else " " for cell in row]
                data.append(row_data)

        return data

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


def extract_csv(csv_path):
    """ This function extracts the text from a CSV file and returns the content as list of lists of strings.
    Each row is represented by a list of strings that contain the cell values.

        Args:
            csv_path(str): Path to the CSV file that is to be converted"""

    name = Path(csv_path).name

    try:
        db = pd.read_csv(csv_path)

        data = []

        for row in db.iterrows():
            row_data = [str(cell) if cell is not None else " " for cell in row][:10]
            data.append(row_data)

        return data

    except FileNotFoundError as e:
        print(f"{_EXCEPTION_MESSAGE_FILE_NOT_FOUND} \n{e}")


def extract_text_from_files(directory_path):
    """ This function goes through all the files in a directory, extracts the text from supported file types as string
     and returns it in a dictionary that had the filename as key and the string content as value.

    Supported file types are: .pdf, .png, .jpg, .giff, .tiff , .bmp, .txt, .docx, .xlsx, .xlsm, .xltx, .xltm, .csv.

           Args:
               directory_path(str): Path to the directory in which the function should run."""

    content_dict = {}

    try:
        for filename in os.listdir(directory_path):

            file_path = os.path.join(directory_path, filename)

            if filename.endswith(".pdf"):
                content_dict[filename] = extract_pdf(file_path)

            elif (filename.endswith(".jpg") or filename.endswith(".jpeg") or filename.endswith(".png") or filename.endswith(".tiff") or
                  filename.endswith(".bmp") or filename.endswith(".gif")):
                content_dict[filename] = extract_image(file_path)

            elif filename.endswith(".txt"):
                content_dict[filename] = extract_txt(file_path)

            elif filename.endswith(".docx"):
                content_dict[filename] = extract_docx(file_path)

            elif (filename.endswith(".xlsx") or filename.endswith(".xlsm") or
                  filename.endswith(".xltx") or filename.endswith(".xltm")):
                content_dict[filename] = extract_excel(file_path)

            elif filename.endswith(".csv"):
                content_dict[filename] = extract_csv(file_path)

            else:
                print(f"Unfortunately the file format for {filename} is not supported. Please try again with a PDF,"
                      " image (.jpg, .png, .giff etc), Excel or Word file.")

    except NotADirectoryError:
        print(f"The input location does not seem to be a directory. Please make sure that the script is "
              f"inside the directory in which you have the files that need converting and try again."
              f"\n{e}")

    except FileNotFoundError as e:
        print(f"The directory that was passed has not been found. Please check the directory "
              f"path is not broken and try again."
              f"\n{e}")

    return content_dict


if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    files_dict = extract_text_from_files(current_dir)
    print(files_dict)
